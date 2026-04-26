"""Shared README/document text extraction helpers.

Primary path uses MarkItDown (Markdown-first conversion). Legacy per-format
extractors remain as fallback for robustness.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path

from bs4 import BeautifulSoup

LOGGER = logging.getLogger(__name__)

_MARKITDOWN = None
_MARKITDOWN_FAILED = False


def _normalize_text(text: str | None) -> str | None:
    if not text:
        return None
    normalized = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    return normalized or None


def _get_markitdown():
    global _MARKITDOWN, _MARKITDOWN_FAILED
    if _MARKITDOWN is not None:
        return _MARKITDOWN
    if _MARKITDOWN_FAILED:
        return None
    try:
        from markitdown import MarkItDown

        _MARKITDOWN = MarkItDown()
        return _MARKITDOWN
    except Exception as exc:  # noqa: BLE001
        _MARKITDOWN_FAILED = True
        LOGGER.warning("MarkItDown unavailable; using fallback extractors: %s", exc)
        return None


def _extract_with_markitdown(raw_bytes: bytes, filename: str) -> str | None:
    converter = _get_markitdown()
    if converter is None:
        return None

    ext = Path(filename).suffix.lower()
    try:
        stream = io.BytesIO(raw_bytes)
        result = converter.convert_stream(stream, file_extension=ext or None)
        markdown = getattr(result, "markdown", None) or getattr(result, "text_content", None)
        return _normalize_text(markdown)
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("MarkItDown extraction failed for %s: %s", filename, exc)
        return None


def _extract_with_fallback(raw_bytes: bytes, filename: str) -> str | None:
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        try:
            import fitz

            pages: list[str] = []
            with fitz.open(stream=raw_bytes, filetype="pdf") as doc:
                for page_obj in doc:
                    text = page_obj.get_text()
                    if text:
                        pages.append(text)
            return _normalize_text("\n\n".join(pages))
        except Exception:
            return None

    if ext == ".docx":
        try:
            import fitz

            pages_d: list[str] = []
            with fitz.open(stream=raw_bytes, filetype="docx") as doc:
                for page_obj in doc:
                    text = page_obj.get_text()
                    if text:
                        pages_d.append(text)
            return _normalize_text("\n\n".join(pages_d))
        except Exception:
            return None

    if ext in (".txt", ".md", ".rst", ""):
        for enc in ("utf-8", "latin-1"):
            try:
                return _normalize_text(raw_bytes.decode(enc))
            except UnicodeDecodeError:
                continue
        return None

    if ext == ".rtf":
        try:
            from striprtf.striprtf import rtf_to_text

            for enc in ("utf-8", "latin-1"):
                try:
                    raw_str = raw_bytes.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return None
            return _normalize_text(rtf_to_text(raw_str, errors="ignore"))
        except Exception:
            return None

    if ext in (".html", ".htm"):
        for enc in ("utf-8", "latin-1"):
            try:
                raw_str = raw_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            return None
        try:
            return _normalize_text(BeautifulSoup(raw_str, "html.parser").get_text("\n", strip=True))
        except Exception:
            return None

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
            parts: list[str] = []
            for sheet in wb.worksheets:
                parts.append(f"[{sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        parts.append("\t".join(cells))
            return _normalize_text("\n".join(parts))
        except Exception:
            return None

    return None


def extract_text_from_bytes(raw_bytes: bytes, filename: str) -> str | None:
    """Extract text from raw file bytes, preferring MarkItDown."""
    primary = _extract_with_markitdown(raw_bytes, filename)
    if primary:
        return primary
    return _extract_with_fallback(raw_bytes, filename)


def extract_text_from_path(file_path: Path) -> str | None:
    """Extract text from a local path, preferring MarkItDown when possible."""
    try:
        raw = file_path.read_bytes()
    except Exception:
        return None
    return extract_text_from_bytes(raw, file_path.name)
